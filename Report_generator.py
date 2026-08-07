# ==========================================================
# IMPORTS
# ==========================================================
import os
import sys
import logging
from datetime import datetime
import pandas as pd
import numpy as np

# Visualization Component Bindings
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Desktop Graphical UI Engine Toolkit
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Excel Native Formatting API Wrappers
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.drawing.image import Image  # Fixed: Imported missing module

# PDF Document Generation Modules
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# ==========================================================
# SYSTEM SETUP & RUNTIME ERROR TRACING
# ==========================================================
logging.basicConfig(
    filename="app_bi_engine.log",
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

# ==========================================================
# CORE STRUCTURAL PIPELINE DATA PREPROCESSING ENGINE
# ==========================================================
class AdvancedDataCleaner:
    @staticmethod
    def deep_preprocess_dataset(df):
        """
        Executes robust multi-stage cleaning over messy data inputs.
        Standardizes types, removes redundant structural information,
        caps mathematical outliers, and returns comprehensive metrics logs.
        """
        metrics = {
            "rows_before": len(df),
            "cols_before": len(df.columns),
            "dupes_removed": 0,
            "missing_filled": 0,
            "outliers_capped": 0,
            "cols_removed": 0,
            "invalid_dates_fixed": 0,
            "mixed_types_coerced": 0
        }
        
        # Phase 1: Deduplication
        df_cleaned = df.copy()
        initial_len = len(df_cleaned)
        df_cleaned = df_cleaned.drop_duplicates()
        metrics["dupes_removed"] = initial_len - len(df_cleaned)

        # Phase 2: Eliminate completely empty tracking columns
        null_cols = [col for col in df_cleaned.columns if df_cleaned[col].isnull().all()]
        if null_cols:
            df_cleaned = df_cleaned.drop(columns=null_cols)
            metrics["cols_removed"] += len(null_cols)
            logging.info(f"CleanEngine pruned empty structural attributes: {null_cols}")

        # Phase 3: Eliminate column attributes with zero data variance (Constant columns)
        constant_cols = [col for col in df_cleaned.columns if df_cleaned[col].nunique(dropna=True) <= 1]
        if constant_cols:
            df_cleaned = df_cleaned.drop(columns=constant_cols)
            metrics["cols_removed"] += len(constant_cols)
            logging.info(f"CleanEngine pruned static non-variance attributes: {constant_cols}")

        # Phase 4: Column-by-column diagnostic data cleansing
        for col in list(df_cleaned.columns):
            # Check for structural Date signatures
            if col.lower() in ['date', 'timestamp', 'created_at', 'order_date']:
                orig_nulls = df_cleaned[col].isnull().sum()
                df_cleaned[col] = pd.to_datetime(df_cleaned[col], errors='coerce')
                new_nulls = df_cleaned[col].isnull().sum()
                metrics["invalid_dates_fixed"] += (new_nulls - orig_nulls)
                continue

            # Standardize mixed object datatypes
            if df_cleaned[col].dtype == 'object':
                # Strip leading and trailing whitespace
                df_cleaned[col] = df_cleaned[col].astype(str).str.strip()
                df_cleaned[col] = df_cleaned[col].replace(['nan', 'NAN', 'Null', 'NULL', ''], np.nan)
                
                # Check if text column consists primarily of numerical data digits
                sample_numeric = pd.to_numeric(df_cleaned[col], errors='coerce').notnull().sum()
                if sample_numeric > (len(df_cleaned) * 0.5):
                    # Coerce and fix mixed data types
                    metrics["mixed_types_coerced"] += 1
                    df_cleaned[col] = pd.to_numeric(df_cleaned[col], errors='coerce')
                else:
                    # Treat as clean categorical text
                    null_count = df_cleaned[col].isnull().sum()
                    if null_count > 0:
                        df_cleaned[col] = df_cleaned[col].fillna("Unknown")
                        metrics["missing_filled"] += null_count
                    continue

            # Handle numeric attributes and cap mathematical outliers
            if pd.api.types.is_numeric_dtype(df_cleaned[col]):
                null_count = df_cleaned[col].isnull().sum()
                if null_count > 0:
                    med_val = df_cleaned[col].median()
                    df_cleaned[col] = df_cleaned[col].fillna(med_val if not pd.isna(med_val) else 0)
                    metrics["missing_filled"] += null_count

                # Bound outliers using the Interquartile Range (IQR) method
                q1 = df_cleaned[col].quantile(0.25)
                q3 = df_cleaned[col].quantile(0.75)
                iqr = q3 - q1
                if iqr > 0:
                    lower_bound = q1 - 1.5 * iqr
                    upper_bound = q3 + 1.5 * iqr
                    
                    outliers_mask = (df_cleaned[col] < lower_bound) | (df_cleaned[col] > upper_bound)
                    metrics["outliers_capped"] += outliers_mask.sum()
                    df_cleaned[col] = np.clip(df_cleaned[col], lower_bound, upper_bound)

        metrics["rows_after"] = len(df_cleaned)
        metrics["cols_after"] = len(df_cleaned.columns)
        
        # Compile summary reports metrics frame
        report_data = {
            "Operational Diagnostic Metric Attribute": [
                "Initial Input Database Rows Volume", "Post-Processed Safe Rows Count",
                "Redundant Row Duplications Cleared", "Missing Structural Cell Data Items Populated",
                "Outlier Skew Values Scaled/Capped", "Redundant Non-Variance Columns Pruned",
                "Corrupted Datetime Attributes Realigned", "Mixed-Type Fields Coerced to Numeric"
            ],
            "Pipeline Computed Value": [
                metrics["rows_before"], metrics["rows_after"],
                metrics["dupes_removed"], metrics["missing_filled"],
                metrics["outliers_capped"], metrics["cols_removed"],
                metrics["invalid_dates_fixed"], metrics["mixed_types_coerced"]
            ]
        }
        report_df = pd.DataFrame(report_data)
        return df_cleaned, report_df


# ==========================================================
# REPORT WRITING ENGINE
# ==========================================================
class AnalyticsEngine:
    @staticmethod
    def extract_functional_columns(df):
        """Categorizes column assignments dynamically into computational vs text groupings."""
        cat_cols = df.select_dtypes(include=["object", "category", "datetime64[ns]"]).columns.tolist()
        num_cols = df.select_dtypes(include=["number"]).columns.tolist()
        return cat_cols, num_cols

    @staticmethod
    def build_pivot_matrix(df, rows, cols, values, operator):
        """Constructs safe multidimensional pivot views based on configuration models."""
        if not rows or not values:
            return pd.DataFrame()
        col_param = cols if (cols and cols != "[ None Mapping ]") else None
        
        try:
            return pd.pivot_table(
                df, index=rows, columns=col_param, values=values,
                aggfunc=operator, fill_value=0
            )
        except Exception as e:
            logging.error(f"Pivot generation calculation failed: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    def render_matplotlib_workspace(pivot, chart_type, val_attr):
        """Compiles clean visual charts based on target data parameters."""
        fig, ax = plt.subplots(figsize=(6, 3.4), facecolor="#1e1e1e")
        ax.set_facecolor("#1e1e1e")
        
        if pivot.empty:
            ax.text(0.5, 0.5, "Data Structural Staging Matrix Empty", color="white", ha='center', va='center')
            return fig

        color_palette = ["#107C41", "#2b579a", "#d24726", "#00828a", "#7e3878", "#ed7d31"]
        
        try:
            if chart_type == "Bar":
                pivot.plot(kind="bar", ax=ax, color=color_palette, zorder=3)
            elif chart_type == "Stacked Bar":
                pivot.plot(kind="bar", stacked=True, ax=ax, color=color_palette, zorder=3)
            elif chart_type == "Horizontal Bar":
                pivot.plot(kind="barh", ax=ax, color=color_palette, zorder=3)
            elif chart_type == "Line":
                pivot.plot(kind="line", marker="o", ax=ax, color=color_palette, zorder=3)
            elif chart_type == "Area":
                pivot.plot(kind="area", alpha=0.4, ax=ax, color=color_palette, zorder=3)
            elif chart_type == "Pie":
                if len(pivot.shape) > 1:
                    pivot.iloc[:, 0].plot(kind="pie", autopct="%1.1f%%", ax=ax, startangle=90)
                else:
                    pivot.plot(kind="pie", autopct="%1.1f%%", ax=ax, startangle=90)
            elif chart_type == "Scatter":
                ax.scatter(range(len(pivot)), pivot.iloc[:, 0] if len(pivot.shape) > 1 else pivot, color="#107C41", zorder=3)
                ax.set_xlabel("Observation Index Item")
            elif chart_type == "Boxplot":
                ax.boxplot(pivot.values, patch_artist=True, boxprops=dict(facecolor="#107C41", color="white"))

            # Refine charts theme parameters to match dark-mode templates UI
            ax.set_title(f"Distribution Metric Analytics Framework: [{val_attr}]", color="white", fontsize=10, fontweight="bold", pad=12)
            ax.tick_params(colors="#b0b0b0", labelsize=8)
            ax.xaxis.label.set_color("#b0b0b0")
            ax.yaxis.label.set_color("#b0b0b0")
            
            if ax.get_legend():
                ax.get_legend().set_frame_on(True)
                ax.get_legend().get_frame().set_facecolor("#2d2d2d")
                for text in ax.get_legend().get_texts():
                    text.set_color("white")
                    text.set_size(7)

            ax.grid(True, color="#333333", linestyle="--", linewidth=0.5, zorder=0)
        except Exception as e:
            ax.clear()
            ax.text(0.5, 0.5, f"Render Matrix Error:\n{str(e)}", color="#ff3333", ha='center', va='center', fontsize=8)
            
        plt.tight_layout()
        return fig


# ==========================================================
# EXCEL STYLING AND FORMATTING ENGINE
# ==========================================================
class ExcelStyles:
    @staticmethod
    def apply_workbook_decorations(filepath):
        """Applies typography formatting, column autofit, headers freezing, and filters."""
        wb = load_workbook(filepath)
        
        # Color definitions
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F9F4", end_color="F2F9F4", fill_type="solid")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Segoe UI", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            
            # Skip advanced styling on raw original logs to avoid accidental compression distortions
            if "Original" in sheetname:
                continue

            # Enable text autofilters and freeze row vectors
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"
            ws.freeze_panes = "A2"

            # Formats body rows cells layouts
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
                for cell in row:
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = font_header
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.font = font_body
                        cell.border = thin_border
                        if row_idx % 2 == 0:
                            cell.fill = zebra_fill
                        
                        # Apply numeric formatting maps based on attribute string features definitions
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'

            # Calculate and configure optimized column text dimensions constraints dynamically
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(filepath)

    @staticmethod
    def compile_dashboard_sheet(writer, clean_df, report_df, pivot_df):
        """Constructs an integrated financial KPI dashboard landing worksheet."""
        wb = writer.book
        ws = wb.create_sheet("Executive Business Dashboard", 0)
        ws.views.sheetView[0].showGridLines = True
        
        # Dashboard Palette Styles Configuration
        card_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
        kpi_title_font = Font(name="Segoe UI", size=9, bold=False, color="A0A0A0")
        kpi_value_font = Font(name="Segoe UI", size=20, bold=True, color="FFFFFF")
        section_font = Font(name="Segoe UI", size=14, bold=True, color="107C41")
        
        # Row 2: Section Title Header Banner
        ws["B2"] = "EXECUTIVE BUSINESS PERFORMANCE DASHBOARD"
        ws["B2"].font = section_font
        
        # Calculate dynamic values for KPI metrics cards
        rows_v = len(clean_df)
        cols_v = len(clean_df.columns)
        dupes_v = int(report_df.iloc[2, 1])
        missing_v = int(report_df.iloc[3, 1])

        kpis = [
            ("TOTAL RECORD ROWS", rows_v, "B", "C"),
            ("TOTAL METRIC COLS", cols_v, "E", "F"),
            ("DUPLICATES CLEARED", dupes_v, "H", "I"),
            ("MISSING CELLS FILLED", missing_v, "K", "L")
        ]

        # Draw structural high-impact summary metric cards layouts
        for title, val, start_col, end_col in kpis:
            # Merge bounds blocks
            ws.merge_cells(f"{start_col}4:{end_col}4")
            ws.merge_cells(f"{start_col}5:{end_col}5")
            
            t_cell = ws[f"{start_col}4"]
            v_cell = ws[f"{start_col}5"]
            
            t_cell.value = title
            t_cell.font = kpi_title_font
            t_cell.fill = card_fill
            t_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            v_cell.value = val
            v_cell.font = kpi_value_font
            v_cell.fill = card_fill
            v_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ingest summary analytics segment view grids directly beneath cards spaces
        ws["B8"] = "Data Cleansing Operational Quality Logs Summary:"
        ws["B8"].font = Font(name="Segoe UI", size=11, bold=True)
        
        for r_idx, row in enumerate(report_df.values, start=9):
            ws[f"B{r_idx}"] = row[0]
            ws[f"C{r_idx}"] = row[1]
            ws[f"C{r_idx}"].number_format = '#,##0'

        # Write current live system slice arrays insights data
        ws["B20"] = "Dynamic Compiled Segment Multi-Aggregation Pivot Matrix Overview:"
        ws["B20"].font = Font(name="Segoe UI", size=11, bold=True)
        
        # Output sub-segment snapshot rows definitions
        pivot_snapshot = pivot_df.head(15)
        for col_idx, col_name in enumerate(pivot_snapshot.columns, start=3):
            ws.cell(row=21, column=col_idx, value=str(col_name)).font = Font(name="Segoe UI", size=10, bold=True)
            
        for r_idx, (idx_lbl, row_vals) in enumerate(zip(pivot_snapshot.index, pivot_snapshot.values), start=22):
            ws.cell(row=r_idx, column=2, value=str(idx_lbl))
            for c_idx, val in enumerate(row_vals, start=3):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.number_format = '#,##0.00'


# ==========================================================
# REFINED APPLICATION INTERFACE
# ==========================================================
class ExcelReportGenerator:
    def __init__(self):
        self.raw_df = None
        self.clean_df = None
        self.report_df = None
        self.plot_canvas = None
        
        self.root = tk.Tk()
        self.root.title("Intelligent Data Preprocessing & BI Analytics WorkBench")
        
        self.root.geometry("1180x750")
        self.root.configure(bg="#121212")
        
        self.build_ui_workspace()
        self.root.mainloop()

    def build_ui_workspace(self):
        # Palette configuration maps setups
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background="#121212")
        style.configure("TLabel", background="#121212", foreground="#E0E0E0", font=("Segoe UI", 10))
        style.configure("Header.TLabel", background="#121212", foreground="#FFFFFF", font=("Segoe UI", 16, "bold"))
        style.configure("TButton", background="#1F1F1F", foreground="#FFFFFF", borderwidth=0, font=("Segoe UI", 10))
        style.map("TButton", background=[("active", "#2D2D2D")])
        style.configure("Action.TButton", background="#107C41", foreground="#FFFFFF", font=("Segoe UI", 10, "bold"))
        style.map("Action.TButton", background=[("active", "#0B592E")])
        
        style.configure("TNotebook", background="#121212", borderwidth=0)
        style.configure("TNotebook.Tab", background="#1F1F1F", foreground="#A0A0A0", padding=[15, 5])
        style.map("TNotebook.Tab", background=[("selected", "#2D2D2D")], foreground=[("selected", "#FFFFFF")])

        # Core operational split regions templates boundaries mapping panels
        self.left_control_panel = ttk.Frame(self.root, padding=20, width=380)
        self.left_control_panel.pack(side=tk.LEFT, fill=tk.Y)
        self.left_control_panel.pack_propagate(False)

        # Right data presentation viewer notebooks containers tabs
        self.right_notebook_panel = ttk.Notebook(self.root)
        self.right_notebook_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Instantiate operational sub-tab container environments panels
        self.tab_preview_raw = ttk.Frame(self.right_notebook_panel, padding=15)
        self.tab_quality_metrics = ttk.Frame(self.right_notebook_panel, padding=15)
        self.tab_pivot_dashboard = ttk.Frame(self.right_notebook_panel, padding=15)

        self.right_notebook_panel.add(self.tab_preview_raw, text=" Staged Raw Source Data View ")
        self.right_notebook_panel.add(self.tab_quality_metrics, text=" Automated Cleaning Diagnostic Logs ")
        self.right_notebook_panel.add(self.tab_pivot_dashboard, text=" Business Intelligence Matrix Dashboard ")

        self.render_left_control_widgets()
        self.render_tab_placeholders()

    # ------------------------------------------------------
    def render_left_control_widgets(self):
        """Constructs and draws left region analytical configurations option control parameters forms."""
        ttk.Label(self.left_control_panel, text="Intelligence Core Engine", style="Header.TLabel").pack(anchor=tk.W, pady=(0, 2))
        ttk.Label(self.left_control_panel, text="Automated extraction and formatting system workspace", 
                  font=("Segoe UI", 8), foreground="#707070").pack(anchor=tk.W, pady=(0, 15))

        ttk.Button(self.left_control_panel, text="📁  Ingest Raw Messy CSV Target File", style="TButton", command=self.execute_ingestion_pipeline).pack(fill=tk.X, pady=6)
        
        self.lbl_status = ttk.Label(self.left_control_panel, text="Status Check: Ingestion Queue Empty", font=("Segoe UI", 9, "italic"), foreground="#D32F2F")
        self.lbl_status.pack(anchor=tk.W, pady=(2, 15))

        # Dynamic parameter configurations runtime tracking values tokens
        self.row_var = tk.StringVar()
        self.col_var = tk.StringVar()
        self.val_var = tk.StringVar()
        self.agg_var = tk.StringVar(value="sum")
        self.chart_var = tk.StringVar(value="Bar")

        # Trace dynamic value overrides to trigger real-time updates to chart views
        for var in [self.row_var, self.col_var, self.val_var, self.agg_var, self.chart_var]:
            var.trace_add("write", self.refresh_live_bi_dashboard_view)

        # Selection element dropdown components interface maps
        ttk.Label(self.left_control_panel, text="Pivot X-Axis Dimension Row Mapping:").pack(anchor=tk.W, pady=(6, 2))
        self.row_cb = ttk.Combobox(self.left_control_panel, textvariable=self.row_var, state="readonly")
        self.row_cb.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(self.left_control_panel, text="Pivot Secondary Dimension Column Breakout:").pack(anchor=tk.W, pady=(6, 2))
        self.col_cb = ttk.Combobox(self.left_control_panel, textvariable=self.col_var, state="readonly")
        self.col_cb.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(self.left_control_panel, text="Pivot Evaluation Target Calculation Field:").pack(anchor=tk.W, pady=(6, 2))
        self.val_cb = ttk.Combobox(self.left_control_panel, textvariable=self.val_var, state="readonly")
        self.val_cb.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(self.left_control_panel, text="Aggregation Operator Evaluation Formula:").pack(anchor=tk.W, pady=(6, 2))
        self.agg_cb = ttk.Combobox(self.left_control_panel, textvariable=self.agg_var, values=["sum", "mean", "count", "max", "min"], state="readonly")
        self.agg_cb.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(self.left_control_panel, text="Dashboard Target Graphical Chart Archetype:").pack(anchor=tk.W, pady=(6, 2))
        self.chart_cb = ttk.Combobox(self.left_control_panel, textvariable=self.chart_var, 
                                     values=["Bar", "Stacked Bar", "Horizontal Bar", "Line", "Area", "Pie", "Scatter", "Boxplot"], state="readonly")
        self.chart_cb.pack(fill=tk.X, pady=(0, 18))

        # Core automation operational compilation dispatchers actions panel items
        ttk.Button(self.left_control_panel, text="📊  Export Advanced Excel Package (.xlsx)", style="Action.TButton", command=self.export_excel_package_suite).pack(fill=tk.X, pady=6)
        ttk.Button(self.left_control_panel, text="📄  Generate Executive Summary Briefing (.pdf)", style="TButton", command=self.export_pdf_briefing_document).pack(fill=tk.X, pady=6)
        
        ttk.Button(self.left_control_panel, text="Shut Down WorkBench Systems", style="TButton", command=self.root.destroy).pack(fill=tk.X, pady=(25, 0))

    # ------------------------------------------------------
    def render_tab_placeholders(self):
        """Constructs fallback states for active reporting dashboards frames."""
        self.lbl_fb1 = tk.Label(self.tab_preview_raw, text="Load raw data stream file inputs to populate layout spreadsheets views grids", font=("Segoe UI", 10, "italic"), bg="#121212", fg="#555555")
        self.lbl_fb1.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        
        self.lbl_fb2 = tk.Label(self.tab_quality_metrics, text="Data cleansing log report items are generated dynamically upon ingestion", font=("Segoe UI", 10, "italic"), bg="#121212", fg="#555555")
        self.lbl_fb2.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Dynamic layout framework for the primary interactive BI split dashboard tab view
        self.p_split_left = tk.Frame(self.tab_pivot_dashboard, bg="#1a1a1a", width=420)
        self.p_split_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        self.p_split_left.pack_propagate(False)

        self.p_split_right = tk.Frame(self.tab_pivot_dashboard, bg="#111111", highlightbackground="#252525", highlightthickness=1)
        self.p_split_right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.lbl_fb3 = tk.Label(self.p_split_left, text="Live Pivot Matrix Tree View Workspace Grid Staging Container", font=("Segoe UI", 9, "italic"), bg="#1a1a1a", fg="#555555", wraplength=250)
        self.lbl_fb3.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        self.lbl_fb4 = tk.Label(self.p_split_right, text="Matplotlib Live Visualization Framework Output Canvas Context", font=("Segoe UI", 9, "italic"), bg="#111111", fg="#555555", wraplength=250)
        self.lbl_fb4.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    # ------------------------------------------------------
    def populate_tree_view_grid(self, container, df):
        """Standard helper module designed to output system data frames visually as responsive grid lists rows."""
        # Clear residual sub-elements inside panel space context
        for widget in container.winfo_children():
            widget.destroy()

        tv = ttk.Treeview(container, columns=list(df.columns), show='headings')
        scrollbar_x = ttk.Scrollbar(container, orient="horizontal", command=tv.xview)
        scrollbar_y = ttk.Scrollbar(container, orient="vertical", command=tv.yview)
        tv.configure(xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set)

        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        tv.pack(fill=tk.BOTH, expand=True)

        for col in df.columns:
            tv.heading(col, text=str(col), anchor=tk.W)
            tv.column(col, width=110, anchor=tk.W)

        # Set maximum visual row output size dynamically to limit screen processing overheads
        for row in df.head(100).values:
            tv.insert('', tk.END, values=[str(x) for x in row])

    # ------------------------------------------------------
    def execute_ingestion_pipeline(self):
        """Handles deep file extraction, executes normalization pipelines, and loads system default dropdown options."""
        path = filedialog.askopenfilename(filetypes=[("Data Files CSV", "*.csv")])
        if not path:
            return
        try:
            self.raw_df = pd.read_csv(path)
            
            # Execute automated structural normalization pipelines overrides
            self.clean_df, self.report_df = AdvancedDataCleaner.deep_preprocess_dataset(self.raw_df)
            
            cat, num = AnalyticsEngine.extract_functional_columns(self.clean_df)
            
            if not cat or not num:
                raise ValueError("Source structural file configuration contains inadequate categorical context attributes matrices fields.")

            # Load choice arrays directly into the UI dropdown variables components
            self.row_cb["values"] = cat
            self.col_cb["values"] = ["[ None Mapping ]"] + cat
            self.val_cb["values"] = num

            # Configure structural baseline defaults automatically based on dataset features
            self.row_cb.current(0)
            self.col_cb.current(0)  # Maps to None Mapping default row layout matrix splitting
            self.val_cb.current(0)

            # Flush out obsolete placeholder fallback labels screens elements cleanly
            if self.lbl_fb1.winfo_exists(): self.lbl_fb1.destroy()
            if self.lbl_fb2.winfo_exists(): self.lbl_fb2.destroy()
            if self.lbl_fb3.winfo_exists(): self.lbl_fb3.destroy()
            if self.lbl_fb4.winfo_exists(): self.lbl_fb4.destroy()

            # Render generated structural information logs frames straight onto tabs views grids
            self.populate_tree_view_grid(self.tab_preview_raw, self.clean_df)
            self.populate_tree_view_grid(self.tab_quality_metrics, self.report_df)

            self.lbl_status.config(text=f"Staging Environment Sync Active: {os.path.basename(path)}", foreground="#107C41")
            
            # Force layout compilation engine update cycle
            self.refresh_live_bi_dashboard_view()
            self.right_notebook_panel.select(2)  # Automatically move focus tabs onto main BI Dashboard workspace layout window
            
            messagebox.showinfo("Ingestion Analytics Sync", "Raw dataset parsed and deep preprocessed successfully!")
        except Exception as e:
            logging.error(f"Ingestion operational failure encountered: {str(e)}")
            messagebox.showerror("Pipeline Sync Interrupted", f"Failed reading structured rows from system metadata:\n{str(e)}")

    # ------------------------------------------------------
    def refresh_live_bi_dashboard_view(self, *args):
        """Dynamically updates the presentation chart grid frames when user-defined dashboard inputs change."""
        if self.clean_df is None:
            return

        try:
            pivot = AnalyticsEngine.build_pivot_matrix(
                self.clean_df, self.row_var.get(), self.col_var.get(), self.val_var.get(), self.agg_var.get()
            )

            # Redraw structural tabular breakdown frames list onto the interactive sub-tab panel view grid area
            pivot_display_df = pivot.reset_index()
            self.populate_tree_view_grid(self.p_split_left, pivot_display_df)

            # Redraw underlying layout configurations charts objects elements safely to clear lingering memory tracking spaces
            if self.plot_canvas:
                self.plot_canvas.get_tk_widget().destroy()

            fig = AnalyticsEngine.render_matplotlib_workspace(pivot, self.chart_var.get(), self.val_var.get())
            
            self.plot_canvas = FigureCanvasTkAgg(fig, master=self.p_split_right)
            self.plot_canvas.draw()
            self.plot_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        except Exception as e:
            logging.warning(f"Live rendering computation cycle skipped safely: {str(e)}")

    # ------------------------------------------------------
    def export_excel_package_suite(self):
        """Compiles datasets sheets, generates KPI dashboard dashboards, styles cells maps, and saves data out to files."""
        if self.clean_df is None:
            messagebox.showerror("Export Interrupted", "Staging matrix data empty. Please ingest a valid file data stream first.")
            return

        try:
            pivot = AnalyticsEngine.build_pivot_matrix(
                self.clean_df, self.row_var.get(), self.col_var.get(), self.val_var.get(), self.agg_var.get()
            )

            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel Workspace Workbooks Portfolio", "*.xlsx")]
            )
            if not save_path:
                return

            temp_img_path = save_path.replace(".xlsx", "_temp_chart_frame.png")

            # Phase 1: Output structural structured database rows records information tables
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                # Compile dashboard landing template view framework
                ExcelStyles.compile_dashboard_sheet(writer, self.clean_df, self.report_df, pivot)
                
                # Write supporting data matrices sheets segments views tabs
                self.clean_df.to_excel(writer, sheet_name="Cleaned Preprocessed Records", index=False)
                self.report_df.to_excel(writer, sheet_name="Data Quality Audit Log", index=False)
                pivot.to_excel(writer, sheet_name="Computed Segment Pivot Matrix")
                self.raw_df.to_excel(writer, sheet_name="Original Dirty Base Ledger Log", index=False)

            # Phase 2: Render analytical charts and inject image artifacts onto dashboards slots
            fig = AnalyticsEngine.render_matplotlib_workspace(pivot, self.chart_var.get(), self.val_var.get())
            fig.savefig(temp_img_path, dpi=130)
            plt.close(fig)

            # Refine cells format overrides constraints properties maps borders line styling elements structures
            ExcelStyles.apply_workbook_decorations(save_path)

            wb = load_workbook(save_path)
            ws = wb["Executive Business Dashboard"]
            if os.path.exists(temp_img_path):
                ws.add_image(Image(temp_img_path), "E8")
                wb.save(save_path)
                os.remove(temp_img_path)

            logging.info(f"Report fully serialized out to physical targeted storage layer location path: {save_path}")
            messagebox.showinfo("Compilation Complete", "Advanced styled analytics business intelligence workbook completely compiled successfully!")

        except Exception as e:
            logging.error(f"Excel generation core execution engine failed to build targets files structures: {str(e)}")
            messagebox.showerror("Export Pipeline Exception", f"Unable to finish compiling target spreadsheet document layout models:\n{str(e)}")

    # ------------------------------------------------------
    def export_pdf_briefing_document(self):
        """Compiles and writes executive print-ready PDF briefing summaries out to file layouts structures."""
        if self.clean_df is None:
            messagebox.showerror("Export Interrupted", "Staging matrix empty. Ingest data target metrics elements first.")
            return

        try:
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".pdf", filetypes=[("Executive Briefing Reports Portfolios", "*.pdf")]
            )
            if not save_path:
                return

            temp_img_path = "temp_pdf_chart_render_mapping.png"
            pivot = AnalyticsEngine.build_pivot_matrix(
                self.clean_df, self.row_var.get(), self.col_var.get(), self.val_var.get(), self.agg_var.get()
            )

            fig = AnalyticsEngine.render_matplotlib_workspace(pivot, self.chart_var.get(), self.val_var.get())
            fig.savefig(temp_img_path, dpi=120)
            plt.close(fig)

            # Initialize ReportLab canvas generator pipelines mapping contexts elements structures
            c = canvas.Canvas(save_path, pagesize=A4)
            width, height = A4

            c.setFont("Helvetica-Bold", 18)
            c.setFillColorRGB(0.06, 0.48, 0.25)  # Hex corporate matching dark green signature accent
            c.drawCentredString(width / 2, height - 50, "Data Analytics Executive Briefing Summary")
            
            c.setStrokeColorRGB(0.8, 0.8, 0.8)
            c.setLineWidth(1)
            c.line(40, height - 65, width - 40, height - 65)

            c.setFont("Helvetica", 10)
            c.setFillColorRGB(0.2, 0.2, 0.2)
            c.drawString(45, height - 90, f"Compilation Execution Date Stamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            c.drawString(45, height - 105, f"Post-Normalized Dataset Inventory Rows: {len(self.clean_df)} | Attributes Columns Size Count: {len(self.clean_df.columns)}")
            c.drawString(45, height - 120, f"Applied Segment Pivot Rule Logic: [Rows: {self.row_var.get()}] * [Values: {self.val_var.get()}] ({self.agg_var.get()})")

            # Draw processing audit text data strings loops logs blocks line statements onto print views panels
            text_block = c.beginText(45, height - 155)
            text_block.setFont("Helvetica-Bold", 12)
            text_block.setFillColorRGB(0, 0, 0)
            text_block.textLine("Automated Pipeline Transformation Summary Metrics Ledger Logs:")
            text_block.setFont("Helvetica", 9)
            text_block.setFillColorRGB(0.3, 0.3, 0.3)
            text_block.textLine("")

            for idx, row in enumerate(self.report_df.values):
                text_block.textLine(f" ->  {row[0]}: {row[1]}")

            c.drawText(text_block)

            # Ingest saved image data metrics snapshots directly inside bounds layout vectors shapes frameworks
            if os.path.exists(temp_img_path):
                c.drawImage(temp_img_path, 45, 60, width=500, height=300)
                c.showPage()
                c.save()
                os.remove(temp_img_path)

            logging.info(f"PDF compiled and closed securely at target path: {save_path}")
            messagebox.showinfo("Export Finalized", "Executive data analytics report file saved successfully as PDF briefing card document portfolio!")
        except Exception as e:
            logging.error(f"PDF creation failure encountered over execution: {str(e)}")
            messagebox.showerror("PDF Export Interrupted Fault", f"System was unable to format reporting page layouts print view blocks:\n{str(e)}")


# ==========================================================
# APPLICATION ENTRY POINT DISPATCHER
# ==========================================================
if __name__ == "__main__":
    ExcelReportGenerator()